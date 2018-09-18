from openpyxl import Workbook


def dump_data(data):
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    for row, rec in enumerate(data, 1):
        ws.cell(row=row, column=1).value = rec.username
        ws.cell(row=row, column=2).value = rec.name
        ws.cell(row=row, column=3).value = rec.about

    wb.save("order.xlsx")
